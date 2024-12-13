import os
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

def get_last_non_empty_row(sheet):
    """
    获取工作表中最后一个非空行的行号
    """
    for row in reversed(range(1, sheet.max_row + 1)):
        if any(cell.value is not None for cell in sheet[row]):
            return row
    return 0

def show():
    # Streamlit界面
    st.title("变压器设备参数导入数据处理工具")

    col1, col2 = st.columns([1, 1])
    with col1:
        old_file = st.file_uploader("上传报送数据表", type=["xlsm"])

    with col2:
        new_file = st.file_uploader("上传新模板表", type=["xlsx"])

    if old_file is not None and new_file is not None:
        # 读取 Excel 文件
        sheet_name1 = '设备线路参数调研'  # 替换为第一个Excel表的Sheet名称
        df1 = pd.read_excel(old_file, sheet_name=sheet_name1)

        # 筛选包含“设备”两个字的行
        filtered_df = df1[df1.apply(lambda row: row.astype(str).str.contains('设备').any(), axis=1)]

        new_df = pd.read_excel(new_file)

        col1, col2 = st.columns([1, 1])
        with col1:
            # 显示旧表和新表
            st.write("报送数据表筛选后的数据：")
            st.write(filtered_df)
        with col2:
            st.write("新模板表：")
            st.write(new_df)

        # 让用户选择要复制的列
        col1, col2 = st.columns([1, 1])
        with col1:
            # 默认选中一些列名
            default_old_columns = ["设备/线路名称", "电压等级", "电流/容量限值（A/MVA）"]  # 这里替换为你希望默认选中的列名
            old_columns = st.multiselect("选择原始数据表中的列", filtered_df.columns, default=default_old_columns)
        with col2:
            # 默认选中一些列名
            default_new_columns = ["变压器绕组名称", "电压等级", "设备限值（MVA)"]  # 这里替换为你希望默认选中的列名
            new_columns = st.multiselect("选择新模板表中的列", new_df.columns, default=default_new_columns)

        # 添加“开始复制”按钮
        if old_columns and new_columns:
            if len(old_columns) == len(new_columns):

                if st.button("开始数据处理"):

                    # 将旧表的选定列数据复制到新表的选定列
                    for old_col, new_col in zip(old_columns, new_columns):
                        new_df[new_col] = filtered_df[old_col].copy()

                    # 定义一个函数来删除后三个字符
                    def remove_last_three_chars(text):
                        if isinstance(text, str) and len(text) > 3:
                            return text[:-3]
                        return text

                    # 应用函数到第二列，并将结果复制到第一列
                    new_df['所属变压器'] = new_df['变压器绕组名称'].apply(remove_last_three_chars)
                    # 提取第二列的最后三个字符并复制到第一列
                    new_df['绕组类型'] = new_df['变压器绕组名称'].str[-3:]

                    column_name = '电压等级'

                    # 定义一个函数来检查并添加“kV”
                    def add_kV(value):
                        # 将值转换为字符串
                        value_str = str(value)
                        if value_str.endswith('kV'):
                            return value_str
                        else:
                            return value_str + 'kV'

                    # 应用函数到指定列
                    new_df[column_name] = new_df[column_name].apply(add_kV)

                    output_file = new_file.name
                    # 检查 output_file 是否存在，如果不存在则复制 new_file 到 output_file
                    if not os.path.exists(output_file):
                        with open(output_file, "wb") as f:
                            f.write(new_file.getbuffer())

                    # 使用 ExcelWriter 追加数据
                    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        # 获取最后一个非空行
                        startrow = writer.sheets['Sheet1'].max_row if writer.sheets['Sheet1'].max_row > 0 else 0
                        new_df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=startrow, header=False)

                    st.success(f"更新后的新表已保存到 {output_file}")
                    df3 = pd.read_excel(output_file, engine='openpyxl')
                    # 显示最终的 DataFrame
                    st.write("处理完成的文件内容：")
                    st.dataframe(df3)
                    with open(output_file, "rb") as f:
                        btn = st.download_button(
                            label="下载更新后的新表",
                            data=f,
                            file_name=output_file,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

            else:
                st.error("原始数据表和新模板表选择的列数必须相同。")
        else:
            st.error("请选择要处理的数据对应的列。")