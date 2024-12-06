import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import streamlit as st

def show():
    # Streamlit界面
    st.title("运行数据处理工具")
    col1, col2 = st.columns([1, 1])
    with col1:
        # 上传第一个 Excel 文件
        input_file = st.file_uploader("上传地调报送数据文件", type=["xlsm"])

    with col2:
        # 上传第二个 Excel 文件
        output_file = st.file_uploader("上传运行数据文件", type=["xlsx"])

    if input_file is not None and output_file is not None:

        # 读取原始 Excel 文件
        # input_file = 'D:/国能日新/地调报送/朝阳/老君堂/老君堂2024.xlsm'
        # output_file = 'D:/国能日新/运行数据/朝阳/老君堂运行数据.xlsx'
        sheet_name = '主变线路运行数据'
        column_name = '名称配置'

        df = pd.read_excel(input_file, sheet_name=sheet_name)
        # 获取从D列开始的所有列数据，并跳过第一行
        columns_to_copy = df.columns[3:]  # D列是第4列，索引为3
        data_to_copy = df.loc[0:, columns_to_copy]  # 从第二行开始复制数据

        # 创建一个新的 DataFrame 来存储转置后的数据
        transposed_data = data_to_copy.T

        # 生成日期列
        start_date = datetime(2024, 5, 1, 0, 0)
        end_date = datetime(2024, 5, 2, 0, 0)
        date_range = pd.date_range(start=start_date, end=end_date, freq='15T')

        # 确保日期列的长度与转置后的数据行数一致
        if len(date_range) > len(transposed_data):
            date_range = date_range[:len(transposed_data)]
        elif len(date_range) < len(transposed_data):
            raise ValueError("日期范围不足以覆盖转置后的数据行数")

        # 将日期列插入到转置后的数据前面
        transposed_data.insert(0, '日期', date_range.strftime('%Y-%m-%d %H:%M'))

        # 获取“名称配置”列的内容，从第0行开始复制
        name_config_data = df[column_name].iloc[0:].reset_index(drop=True).tolist()  # 从第0行开始复制，并转换为列表

        # 确保 name_config_data 的长度与 transposed_data 的列数一致
        if len(name_config_data) < len(transposed_data.columns) - 1:
            name_config_data = name_config_data + [''] * (len(transposed_data.columns) - 1 - len(name_config_data))
        elif len(name_config_data) > len(transposed_data.columns) - 1:
            name_config_data = name_config_data[:len(transposed_data.columns) - 1]

        # 创建新增行数据
        new_row = ['日期'] + name_config_data

        # 创建一个新的 DataFrame 来存储增加一行后的数据
        new_df = pd.DataFrame([new_row], columns=transposed_data.columns)

        # 将新增行插入到转置后的数据前面
        final_data = pd.concat([new_df, transposed_data], ignore_index=True)

        # 创建一个新的 Excel 工作簿
        wb = Workbook()
        ws = wb.active

        # 将 DataFrame 写入工作表
        for r in dataframe_to_rows(final_data, index=False, header=False):
            ws.append(r)

        # 设置数据居中
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 自动调整列宽，只考虑第一行的内容长度
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            # 获取第一行的内容长度
            first_row_length = len(str(ws.cell(row=1, column=col[0].column).value))
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # 比较第一行的内容长度和数据的最大长度
            max_length = max(max_length, first_row_length)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 保存工作簿
        wb.save(output_file)

        st.success(f"数据已成功转置并保存到 {output_file}")
        st.divider()
        df3 = pd.read_excel(output_file)
        # 显示最终的 DataFrame
        st.write("处理后的数据文件内容：")
        st.write(df3)
